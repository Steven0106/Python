from openpyxl import load_workbook
from openpyxl import Workbook
from glob import *
import os

# delete current print
try:
    os.remove("group_python.xlsx")
except:
     pass


Myfiles = [i for i in glob('*.xlsx')]
# print(Myfiles)
total_student = []
# save info per files
for item in Myfiles:
    my_workbook = load_workbook(item, data_only = True)
    my_worksheet = my_workbook['Sheet1']
    my_list = []
    my_list.append(my_worksheet['A2'].value)
    my_list.append(my_worksheet['B2'].value)
    my_list.append(my_worksheet['C2'].value)
    my_list.append(my_worksheet['D2'].value)
    total_student.append(my_list)
    print(my_list)



# all dict for student
total_student_by_group = {}
for i in range (10):
    total_student_by_group["group"+str(i+1)] = {}
print(total_student_by_group)

# group compare dict
assign_dict = {}
for i in range (10):
    assign_dict[i+1] = "group"+str(i+1)
print(assign_dict)

# student per cycle
student_number_tracker = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

print(total_student)
# sort stud by group
for item in total_student:
    group = item[2]
    group_name = assign_dict[group]
    total_student_by_group[group_name]["student_"+str(student_number_tracker[group-1]+1)] = item
    student_number_tracker[group-1] += 1

print(total_student_by_group)
print(student_number_tracker)
# my_dict = {}
# my_dict["stu_ID"] = my_worksheet['A2'].value
# my_dict["name"] = my_worksheet['B2'].value
# my_dict["group_id"] = my_worksheet['C2'].value
# my_dict["git"] = my_worksheet['D2'].value
# print(my_dict)
# print(my_worksheet['A2'].value)
# print(my_worksheet['B2'].value)
# print(my_worksheet['C2'].value)
# print(my_worksheet['D2'].value)

####
my_writing_wb = Workbook()

# make 10 sheet
for i in range(10):
    write_ws = my_writing_wb.create_sheet("jo"+str(i+1))

# part for print in excel
for i in range(10):
    # get current using worksheet
    load_ws = my_writing_wb["jo"+str(i+1)]
    # add header
    load_ws.append(["stu_num", "name", "group_id", "git"])

    # save group temp
    tempList = list(total_student_by_group["group"+str(i+1)].values())
    print(tempList)
    for j in range(4):
        try:
            load_ws.append(tempList[j])
        except:
            pass


my_writing_wb.remove(my_writing_wb['Sheet'])

# save in same folder with py
my_writing_wb.save("group_python.xlsx")
